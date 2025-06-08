from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl

from administrator.serializers import ResultSerializer
from .models import AcademicSession, Result, Student, Subject, SchoolProfile, Term, TermTotalMark
from rest_framework.parsers import MultiPartParser, FormParser

class SubjectExcelExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, student_id):
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student"}, status=404)
        
        school = SchoolProfile.objects.filter(user=request.user).first()
        if not school:
            return Response({"error": "School profile not found."}, status=404)
        
        session = AcademicSession.objects.filter(school = school, is_current=True).first()
        if not session:
            return Response({"error": "session not set"}, status=404)
        
        term = Term.objects.filter(session=session, is_current=True).first()
        if not term:
            return Response({"error": "term not set"}, status=404)

        results = Result.objects.filter(student=student, term=term, session=session)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        headers = ['Subject', 'CA1', 'CA2', 'CA3', 'Exam']
        ws.append(headers)

        if results.exists():
            result_subject_names = set(result.subjects.strip() for result in results if result.subjects)

            for result in results:
                ws.append([
                    result.subjects,
                    result.first_test or '',
                    result.second_test or '',
                    result.third_test or '',
                    result.exam or ''
                ])

            # Append subjects not in result
            school_subjects = Subject.objects.filter(school=school).order_by('name')
            for subject in school_subjects:
                if subject.name not in result_subject_names:
                    ws.append([subject.name, '', '', '', ''])
        else:
            # No results; generate blank template
            subjects = Subject.objects.filter(school=school).order_by('name')
            for subject in subjects:
                ws.append([subject.name, '', '', '', ''])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"{student.name.replace(' ', '_')}_result_template.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response




class UploadStudentResultPreviewView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, student_id):
        excel_file = request.FILES.get('file')
        if not all([student_id, excel_file]):
            return Response({'detail': 'Missing required field or file'}, status=400)

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student"}, status=404)

        school = SchoolProfile.objects.filter(user=request.user).first()
        if not school:
            return Response({"error": "School profile not found."}, status=404)

        session = AcademicSession.objects.filter(school=school, is_current=True).first()
        if not session:
            return Response({"error": "Session not set"}, status=404)

        term = Term.objects.filter(session=session, is_current=True).first()
        if not term:
            return Response({"error": "Term not set"}, status=404)

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected_headers = ['Subject', 'CA1', 'CA2', 'CA3', 'Exam']
        if headers != expected_headers:
            return Response({'detail': 'Invalid file format. Expected headers: ' + ', '.join(expected_headers)}, status=400)

        db_subjects = set(
            Subject.objects.filter(school=school).values_list('name', flat=True)
        )


        valid_rows = []
        skipped_rows = []
        uploaded_subjects = set()

        def is_valid_score(value):
            try:
                if value is None:
                    return False
                int(value)
                return True
            except (ValueError, TypeError):
                return False

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            subject_name = str(row[0]).strip() if row[0] else None
            ca1, ca2, ca3, exam = row[1], row[2], row[3], row[4]

            # Check subject name present and valid
            if not subject_name or subject_name not in db_subjects:
                skipped_rows.append({'row': idx, 'reason': 'Invalid or missing subject name', 'data': row})
                continue

            # Check all scores are valid (not None and intable)
            if not all(is_valid_score(score) for score in (ca1, ca2, ca3, exam)):
                skipped_rows.append({'row': idx, 'reason': 'One or more scores missing or invalid', 'data': row})
                continue

            uploaded_subjects.add(subject_name)
            valid_rows.append({
                'subject': subject_name,
                'ca1': int(ca1),
                'ca2': int(ca2),
                'ca3': int(ca3),
                'exam': int(exam)
            })

        invalid_subjects = list(uploaded_subjects - db_subjects)
        if invalid_subjects:
            return Response({
                'detail': 'Some subjects are invalid or do not match exactly with the database.',
                'invalid_subjects': invalid_subjects,
                'skipped_rows': skipped_rows
            }, status=400)

        return Response({
            "session": f"{session.name} | {term.name}",
            "student": student.name,
            "student_id": student.id,
            'detail': 'Preview of uploaded results',
            'valid_rows': valid_rows,
            'skipped_rows': skipped_rows
        }, status=200)



class ConfirmUploadStudentResultView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data.get('results', [])
        student_id = request.data.get('student_id')
        teacher_comment = request.data.get("teacher_comment")
        principal_comment = request.data.get("principal_comment")
        
        if not data:
            return Response({'detail': 'No results provided'}, status=400)

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student"}, status=404)

        school = SchoolProfile.objects.filter(user=request.user).first()
        if not school:
            return Response({"error": "School profile not found."}, status=404)

        session = AcademicSession.objects.filter(school=school, is_current=True).first()
        if not session:
            return Response({"error": "Session not set"}, status=404)

        term = Term.objects.filter(session=session, is_current=True).first()
        if not term:
            return Response({"error": "Term not set"}, status=404)

        for item in data:
            Result.objects.update_or_create(
                student=student,
                term=term,
                session=session,
                subjects=item['subject'],
                defaults={
                    'first_test': item.get('ca1'),
                    'second_test': item.get('ca2'),
                    'third_test': item.get('ca3'),
                    'exam': item.get('exam')
                }
            )
            
        resultSummary = TermTotalMark.objects.filter(student=student, term=term, session=session).first()
        if resultSummary:
            resultSummary.principal_comment = principal_comment
            resultSummary.teacher_comment = teacher_comment
            resultSummary.save()
            

        return Response({'detail': '✅ Results saved successfully.'}, status=201)




class ResetStudentResultView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, student_id):
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student"}, status=404)
        
        school = SchoolProfile.objects.filter(user=request.user).first()
        if not school:
            return Response({"error": "School profile not found."}, status=404)
        
        session = AcademicSession.objects.filter(school = school, is_current=True).first()
        if not session:
            return Response({"error": "session not set"}, status=404)
        
        term = Term.objects.filter(session=session, is_current=True).first()
        if not term:
            return Response({"error": "term not set"}, status=404)
        
        results = Result.objects.filter(student=student, term=term, session=session)
        if not results.exists():
            return Response({"error": "No results found for this student in the current term and session."}, status=404)

        results.delete()
        return Response({'detail': '✅ Results reset successfully.'}, status=200)
    
    
from django.db.models import Sum, Count, F
from django.db.models.functions import Coalesce

def ordinal(n):
    return f"{n}{'tsnrhtdd'[(n//10%10!=1)*(n%10<4)*n%10::4]}"


class ShowStudentResultView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, student_id):
        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return Response({"error": "Invalid student"}, status=404)
        
        school = SchoolProfile.objects.filter(user=request.user).first()
        if not school:
            return Response({"error": "School profile not found."}, status=404)
        
        admin_user = school.user.filter(is_admin=True).first()
        if not admin_user:
            return Response({"error": "No admin user found for this school."}, status=404)
        
        session = AcademicSession.objects.filter(school = school, is_current=True).first()
        if not session:
            return Response({"error": "session not set"}, status=404)
        
        term = Term.objects.filter(session=session, is_current=True).first()
        if not term:
            return Response({"error": "term not set"}, status=404)
        
        results = Result.objects.filter(student=student, term=term, session=session)
        if not results.exists():
            return Response({"error": "No results found for this student in the current term and session."}, status=404)
        results_serializer = ResultSerializer(results, many=True)
        
        # Comments
        resultSummary = TermTotalMark.objects.filter(student=student, term=term, session=session).first()        
        
        # Total score for the student
        total_score = results.aggregate(total=Coalesce(Sum('total_score'), 0))['total']
        subjects_count = results.count()
        total_possible = subjects_count * 100
        average_score = round((total_score / total_possible) * 100, 2) if total_possible else 0
        
        
        # Class-level stats
        class_students = Student.objects.filter(class_level=student.class_level)
        class_results = Result.objects.filter(
            student__in=class_students,
            session=session,
            term=term
        )
        
        
        # Compute each student's total
        student_scores = (
            class_results.values('student_id')
            .annotate(total=Coalesce(Sum('total_score'), 0))
            .order_by('-total')
        )
        
        
        # Calculate position
        student_position = 1
        for idx, data in enumerate(student_scores, start=1):
            if data['student_id'] == student.id:
                student_position = idx
                break
            
        # Class average
        class_total = class_results.aggregate(class_total=Coalesce(Sum('total_score'), 0))['class_total']
        total_subjects = class_results.count()
        student_total_count = class_students.count()
        class_avg = round((class_total / (student_total_count * 100 * subjects_count)) * 100, 2) if student_total_count and subjects_count else 0

        
        
        return Response({
            "school_info": {
                "school_name": school.school_name,
                "location": school.school_address or "Not Set",
                "phone": admin_user.phone ,
                "email":  admin_user.email
            },
            "academic_sessions" : {
                "session": session.name,
                "term": term.name,
            },
            "student" : {
                "student_name" : student.name,
                "other_info" : student.other_info,
                "class" : student.class_level.name
            },
            "results" : results_serializer.data,
            "performance_summary": {
                "total_score": total_score,
                "out_of": total_possible,
                "average_score": f"{average_score}%",
                "class_average": f"{class_avg}%",
                "position": ordinal(student_position),
                "out_of_students": student_total_count
            },
            "comments":{
                "principal_comment": resultSummary.principal_comment if resultSummary and resultSummary.principal_comment else "Not Set",
                "teacher_comment": resultSummary.teacher_comment if resultSummary and resultSummary.teacher_comment else "Not Set",

            }
        })